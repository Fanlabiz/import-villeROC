import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="RoC - Import Ville", page_icon="🏙️", layout="centered")
st.title("🏙️ Rise of Cultures — Import Ville")
st.caption("Convertit le CSV exporté depuis le jeu en fichier Excel pour l'optimiseur.")

# ── Catalogue dimensions (largeur, hauteur) ──
# Ordre important : les clés plus spécifiques d'abord
DIMS = {
    # Culture Sites
    'CultureSite_Large':     (3, 3),
    'CultureSite_Luxurious': (2, 2),
    'CultureSite_Moderate':  (2, 2),
    'CultureSite_Compact':   (2, 1),
    'CultureSite_Small':     (1, 2),
    'CultureSite_Little':    (1, 1),
    # Maisons
    'Home_Large':    (3, 2),
    'Home_Average':  (2, 2),
    'Home_Small':    (2, 2),
    'Home_Medium':   (2, 2),
    # Fermes — corrections
    'Farm_Domestic': (4, 4),
    'Farm_Premium':  (3, 3),
    'Farm_Rural':    (3, 2),
    'Farm_Pastoral': (3, 3),
    # Casernes
    'Barracks_Siege':         (5, 6),
    'Barracks_HeavyInfantry': (3, 3),
    'Barracks_Infantry':      (3, 3),
    'Barracks_Ranged':        (3, 3),
    'Barracks_Cavalry':       (3, 3),
    # Workshops
    'Workshop_Alchemist':    (3, 3),
    'Workshop_Glassblower':  (3, 3),
    'Workshop_Jeweler':      (3, 3),
    'Workshop_Carpenter':    (3, 3),
    'Workshop_Scribe':       (3, 3),
    'Workshop_SpiceMerchant':(3, 3),
    'Workshop_Coffee':       (2, 2),
    'Workshop_Incense':      (2, 2),
    'Workshop_OilLamp':      (2, 2),
    'Workshop_Carpet':       (2, 2),
    # Génériques
    'CityHall':    (4, 4),
    # Bâtiments évolutifs (dimensions spécifiques connues)
    'Evolving_DraculaCastle': (4, 3),
    'Evolving_CryptOfTheCount': (4, 3),
    'Evolving_MadScientistsLab': (4, 3),
    'Evolving_HotAirBalloon': (2, 2),
    'Evolving_SleighStation': (2, 2),
    'Evolving_WinterMarket': (3, 3),
    'Evolving_Bakery': (2, 2),
    'Evolving_RoyalTemple': (3, 3),
    'Evolving_AztecMainTemple': (4, 4),
    'Evolving_Aqueduct': (3, 3),
    'Evolving_MotherTree': (3, 3),
    'Evolving_TravellingYurt': (2, 2),
    'Evolving_YurtOfTheKhan': (3, 3),
    'Evolving_MongolianFeast': (3, 3),
    'Evolving_CelticArch': (3, 3),
    'Evolving_GrandSmithy': (3, 3),
    'Evolving_CelticBroch': (3, 3),
    'Evolving_GriotCourt': (3, 3),
    'Evolving_Madrasa': (4, 3),
    'Evolving_DovecoteTower': (2, 3),
    'Evolving_DrumTower': (3, 3),
    'Evolving_FountainOfYouth': (3, 3),
    'Evolving_PirateFortress': (4, 4),
    'Evolving_TreasureWreck': (3, 3),
    'Evolving_ElysianField': (3, 3),
    'Evolving_GreatGarden': (3, 3),
    'Evolving_Conservatory': (3, 3),
    'Evolving_AncientLibrary': (3, 3),
    'Evolving_Hydra': (3, 3),
    'Evolving_TrojanHorse': (3, 3),
    'Evolving_Exhibition': (4, 3),
    'Evolving_NaturalHistoryMuseum': (4, 3),
    'Evolving_ShrineOfReflection': (3, 3),
    'Evolving_ShirazWineHouse': (3, 3),
    'Evolving_MosaicBath': (3, 3),
    'Evolving': (3, 3),  # fallback générique
    'Collectable': (2, 2),
    'Wonder':      (4, 4),
    'Harbor':      (3, 3),
    # Arabia
    'Irrigation_Noria':  (2, 2),
    'Irrigation_Large':  (3, 2),
    'Irrigation_Medium': (2, 2),
    'Irrigation_Small':  (1, 2),
    'Irrigation_Oasis':  (3, 3),
    'Merchant_Average':  (2, 2),
    'CamelFarm_Average': (3, 3),
    # Fallback
    'DEFAULT': (2, 2),
}

CATS = {
    'CultureSite': 'Culturel',
    'Barracks':    'Caserne',
    'Farm':        'Production',
    'Home':        'Habitation',
    'CityHall':    'Neutre',
    'Evolving':    'Producteur',
    'Collectable': 'Neutre',
    'Wonder':      'Neutre',
    'Irrigation':  'Neutre',
    'Merchant':    'Production',
    'CamelFarm':   'Production',
    'Workshop':    'Production',
    'Harbor':      'Neutre',
    'DEFAULT':     'Neutre',
}

# Valeurs culture et rayonnement lues depuis le CSV (extraites du GameDesignResponse)
CULTURE_VAL = {}
CULTURE_RANGE = {}

# Seuils de culture pour les boosts de production (depuis f11 du BuildingDefinitionDTO)
# Format : { 'Building_xxx': {25: seuil_25pct, 50: seuil_50pct, 100: seuil_100pct} }
DIMS_THRESHOLDS = {}

# Formules dynamiques de culture pour les bâtiments Evolving (depuis DynamicFloatValueDefinitionDTO)
# Format: { 'motif_nom': { 'EreJoueur': lambda niveau: valeur } }
# L'ère du joueur est déduite du niveau f18 dans le CSV (mapping index → ère)
_ERES_PAR_INDEX = {
    1:'StoneAge',2:'BronzeAge',3:'MinoanEra',4:'ClassicGreece',5:'EarlyRome',
    6:'RomanEmpire',7:'ByzantineEra',8:'AgeOfTheFranks',9:'FeudalAge',
    10:'IberianEra',11:'KingdomOfSicily',12:'HighMiddleAges',13:'EarlyGothicEra',
    14:'LateGothicEra',15:'LateGothicEra',
}
_CULTURE_FORMULAS = {
    'Aqueduct': {
        'StoneAge':lambda l:(l+18)*2,'BronzeAge':lambda l:(l+18)*3,
        'MinoanEra':lambda l:(l+18)*4,'ClassicGreece':lambda l:(l+18)*5,
        'EarlyRome':lambda l:(l+18)*6,'RomanEmpire':lambda l:(l+18)*7,
        'ByzantineEra':lambda l:(l+18)*8,'AgeOfTheFranks':lambda l:(l+18)*9,
        'FeudalAge':lambda l:(l+18)*10,'IberianEra':lambda l:(l+18)*11,
        'KingdomOfSicily':lambda l:(l+18)*12,'HighMiddleAges':lambda l:(l+18)*13,
        'EarlyGothicEra':lambda l:(l+18)*14,'LateGothicEra':lambda l:(l+18)*15,
    },
    'GreatGarden': {
        'StoneAge':lambda l:l*3.5+89,'BronzeAge':lambda l:l*4+114,
        'MinoanEra':lambda l:l*4.5+139,'ClassicGreece':lambda l:l*6+127,
        'EarlyRome':lambda l:l*6.5+152,'RomanEmpire':lambda l:l*7+177,
        'ByzantineEra':lambda l:l*8+184,'AgeOfTheFranks':lambda l:l*8.5+209,
        'FeudalAge':lambda l:l*9+234,'IberianEra':lambda l:l*10+241,
        'KingdomOfSicily':lambda l:l*11+265,'HighMiddleAges':lambda l:l*12+288,
        'EarlyGothicEra':lambda l:l*13+320,'LateGothicEra':lambda l:l*14+339,
    },
}
_RANGE_FORMULAS = {
    # Rayonnement selon niveau pour les Evolving
    'Aqueduct': lambda l: 1 if l < 30 else (2 if l < 60 else 3),
    'GreatGarden': lambda l: 1 if l < 12 else (2 if l < 25 else (3 if l < 38 else 4)),
}

def compute_evolving_culture(nom_complet, niveau, ere_index):
    """Calcule la culture d'un Evolving depuis les formules du GameDesignResponse."""
    ere = _ERES_PAR_INDEX.get(ere_index, 'LateGothicEra')
    for key, formulas in _CULTURE_FORMULAS.items():
        if key in nom_complet:
            fn = formulas.get(ere, formulas.get('LateGothicEra'))
            if fn:
                return int(round(fn(niveau)))
    return 0

def compute_evolving_range(nom_complet, niveau):
    for key, fn in _RANGE_FORMULAS.items():
        if key in nom_complet:
            return fn(niveau)
    return 0

def _load_thresholds_from_catalog(catalog_dict):
    """Peuple DIMS_THRESHOLDS depuis le catalogue déjà parsé."""
    global DIMS_THRESHOLDS
    DIMS_THRESHOLDS = {}
    for bname, info in catalog_dict.items():
        t = info.get('thresholds', {})
        if t:
            DIMS_THRESHOLDS[bname] = t

def _get_threshold(full_name, pct):
    """Retourne le seuil de culture pour un boost donné (25, 50 ou 100)."""
    t = DIMS_THRESHOLDS.get(full_name)
    if not t:
        # Cherche sans numéro de niveau
        base = _re.sub(r'_\d+$', '', full_name)
        for k, v in DIMS_THRESHOLDS.items():
            if _re.sub(r'_\d+$', '', k) == base:
                t = v; break
    return t.get(pct, '') if t else ''


import re as _re

# 3 couleurs : orange (culture), vert (producteur), gris (neutre)
CAT_COLORS = {
    'Culturel':    'FFB347',
    'Caserne':     'A8D5A2',
    'Production':  'A8D5A2',
    'Habitation':  'A8D5A2',
    'Producteur':  'A8D5A2',
    'Neutre':      'DCDCDC',
}

# Mots-clés identifiant un bâtiment culturel "réel"
# Tous les Collectable et Evolving produisent de la culture
_CULTURE_KEYS = ['CultureSite', 'Evolving', 'Collectable', 'CityHall']

def get_key(name, mapping):
    for k in mapping:
        if k == 'DEFAULT': continue
        if k in name: return k
    return 'DEFAULT'

def get_cat(name):
    return CATS.get(get_key(name, CATS), 'Neutre')

def is_real_culture(name, range_val):
    """True si le bâtiment produit de la culture (CultureSite, Evolving, Collectable, CityHall)."""
    return any(k in name for k in _CULTURE_KEYS)

def get_color(name, range_val):
    if is_real_culture(name, range_val):
        return 'FFB347'   # orange
    cat = get_cat(name)
    if cat in ('Habitation', 'Production', 'Producteur', 'Caserne'):
        return 'A8D5A2'   # vert
    return 'DCDCDC'       # gris

# Préfixes d'ères/events à supprimer du nom affiché
_ERA_PREFIXES = [
    'EventHalloween2021', 'EventHalloween2022', 'EventHalloween',
    'EventWinter2021', 'EventWinter2022', 'EventWinter',
    'EventGreek2023', 'EventGreek',
    'EventMongols2022', 'EventMongols2023', 'EventMongol',
    'EventCelts2022', 'EventCelts2023', 'EventCeltic',
    'EventMaliEmpire2022', 'EventMaliEmpire2023', 'EventMaliEmpire',
    'EventThai2022', 'EventThai',
    'EventAztec', 'EventPersian2023', 'EventPersian',
    'EventPolynesia', 'EventHercules2021', 'EventHercules2022',
    'EventWorldFair', 'TreasureHunt', 'SeasonPass', 'PlayerEncounters',
    'RoCBirthday', 'MinoanEra', 'BronzeAge', 'StoneAge', 'ClassicGreece',
    'EarlyRome', 'RomanEmpire', 'ByzantineEra', 'AgeOfTheFranks',
    'FeudalAge', 'IberianEra', 'KingdomOfSicily', 'HighMiddleAges',
    'EarlyGothicEra', 'LateGothicEra', 'DynamicAge', 'Harbor',
    'Evolving', 'Collectable', 'City',
]

def clean_name(full_name):
    """Retourne le nom court lisible d'un bâtiment."""
    n = full_name.replace('Building_', '')
    for p in _ERA_PREFIXES:
        n = n.replace(p + '_', '')
    n = _re.sub(r'_\d+$', '', n)
    return n

def get_level(full_name):
    m = _re.search(r'_(\d+)$', full_name)
    return int(m.group(1)) if m else 1

def short_name(name):
    return clean_name(name)

def build_excel(df, ville):
    rows_df = df[df['Ville'] == ville].copy()
    if rows_df.empty:
        return None

    # Calcule les vraies limites du terrain
    # En tenant compte des dimensions des bâtiments
    max_r_used = 0
    max_c_used = 0
    for _, b in rows_df.iterrows():
        bkey = get_key(b['Nom_complet'], DIMS)
        w, h = DIMS[bkey]
        max_r_used = max(max_r_used, int(b['Ligne']) + h - 1)
        max_c_used = max(max_c_used, int(b['Colonne']) + w - 1)

    # Terrain avec marge de 3 cases
    n_rows = max_r_used + 4
    n_cols = max_c_used + 4

    wb = Workbook()

    # ── Styles ──
    thin      = Side(style='thin',   color='AAAAAA')
    thick     = Side(style='medium', color='333333')
    brd_cell  = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_inner = Border(left=thin, right=thin, top=thin, bottom=thin)

    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center')

    hdr_fill = PatternFill('solid', start_color='366092')
    hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    free_fill   = PatternFill('solid', start_color='F5F5F5')
    border_fill = PatternFill('solid', start_color='D0D0D0')

    # ── Onglet Terrain ──
    ws_t = wb.active
    ws_t.title = 'Terrain'

    # Après CCW: lignes Excel = colonnes jeu (n_cols lignes), colonnes Excel = lignes jeu (n_rows colonnes)
    ws_t.cell(1, 1).value = ''
    for c in range(n_rows):          # colonnes Excel correspondent aux lignes jeu
        cell = ws_t.cell(1, c + 2)
        cell.value = n_rows - 1 - c  # ligne 0 jeu à droite Excel
        cell.font = Font(bold=True, name='Arial', size=8)
        cell.alignment = ctr
        cell.fill = PatternFill('solid', start_color='E8E8E8')

    for r in range(n_cols):          # lignes Excel correspondent aux colonnes jeu
        cell = ws_t.cell(r + 2, 1)
        cell.value = n_cols - 1 - r  # colonne 0 jeu en bas Excel
        cell.font = Font(bold=True, name='Arial', size=8)
        cell.alignment = ctr
        cell.fill = PatternFill('solid', start_color='E8E8E8')

    # Grille de fond (n_cols lignes × n_rows colonnes après CCW)
    for r in range(n_cols):
        for c in range(n_rows):
            cell = ws_t.cell(r + 2, c + 2)
            cell.fill = free_fill
            cell.border = brd_cell

    # Détermine l'orientation correcte de chaque bâtiment non carré
    # par analyse d'adjacence : la bonne orientation est celle qui a des voisins
    # exactement adjacents (pas de gap, pas de chevauchement).

    # Orientation : lue directement depuis le CSV (colonne Rotation = f9 du CityDTO)
    # f9=1 → dimensions échangées par rapport au GameDesignResponse
    # Si la colonne Rotation n'existe pas (ancien CSV), on utilise l'algorithme d'adjacence
    has_rotation_col = 'Rotation' in rows_df.columns

    all_buildings = []
    for _, b in rows_df.iterrows():
        w0 = int(b['Largeur']) if pd.notna(b.get('Largeur')) and int(b.get('Largeur',0)) > 0 else 2
        h0 = int(b['Hauteur']) if pd.notna(b.get('Hauteur')) and int(b.get('Hauteur',0)) > 0 else 2
        all_buildings.append({
            'nom': b['Nom_complet'], 'col': int(b['Colonne']), 'row': int(b['Ligne']),
            'w': w0, 'h': h0
        })

    def best_orientation_adjacency(col, row, w, h, buildings):
        """Fallback : orientation par adjacence (ancien CSV sans colonne Rotation)."""
        if w == h: return w, h
        def adj4(c0, r0, w, h):
            right=below=left=above=0
            for o in buildings:
                oc,or_,ow,oh = o['col'],o['row'],o['w'],o['h']
                if oc==c0+w  and or_<r0+h and or_+oh>r0: right+=1
                if or_==r0+h and oc<c0+w  and oc+ow>c0: below+=1
                if c0==oc+ow and or_<r0+h and or_+oh>r0: left+=1
                if r0==or_+oh and oc<c0+w and oc+ow>c0: above+=1
            return right,below,left,above
        r1,b1,l1,a1 = adj4(col,row,w,h)
        r2,b2,l2,a2 = adj4(col,row,h,w)
        s1,s2 = r1+b1+l1+a1, r2+b2+l2+a2
        if s2>s1: return h,w
        if s1>s2: return w,h
        vert1,vert2 = b1+a1, b2+a2
        if vert2>vert1: return h,w
        if vert1>vert2: return w,h
        return min(w,h),max(w,h)

    # Applique l'orientation
    oriented = []
    for idx, b in enumerate(all_buildings):
        row_b = rows_df.iloc[idx]
        if has_rotation_col:
            # Dimensions déjà corrigées dans le CSV par le userscript (f9 appliqué)
            w, h = b['w'], b['h']
        else:
            w, h = best_orientation_adjacency(b['col'], b['row'], b['w'], b['h'], all_buildings)
        b = dict(b); b['w'] = w; b['h'] = h
        oriented.append(b)

    # Place les bâtiments sans chevauchement (tri par surface décroissante)
    placed = []
    occupied = set()

    def try_place(r0, c0, w, h):
        cells = [(r0+dr, c0+dc) for dr in range(h) for dc in range(w)]
        if any(cell in occupied for cell in cells):
            return False
        for cell in cells: occupied.add(cell)
        return True

    oriented_sorted = sorted(oriented, key=lambda b: b['w']*b['h'], reverse=True)

    # Construit un index nom → (culture, rayonnement) depuis rows_df
    cult_map = {}
    for _, b in rows_df.iterrows():
        cult_map[b['Nom_complet'] + '|' + str(int(b['Colonne'])) + '|' + str(int(b['Ligne']))] = (
            int(b['Culture']) if pd.notna(b.get('Culture')) else 0,
            int(b['Rayonnement']) if pd.notna(b.get('Rayonnement')) else 0,
        )

    for b in oriented_sorted:
        r0, c0, w, h = b['row'], b['col'], b['w'], b['h']
        nom = b['nom']
        key = nom + '|' + str(c0) + '|' + str(r0)
        cult_val, ray_val = cult_map.get(key, (0, 0))
        lbl   = short_name(nom)
        color = get_color(nom, ray_val)
        if try_place(r0, c0, w, h):
            placed.append((r0, c0, w, h, lbl, color))
        elif w != h and try_place(r0, c0, h, w):
            placed.append((r0, c0, h, w, lbl, color))

    # Rotation CCW (¼ de tour antihoraire) :
    # (col_jeu, row_jeu) → dans la grille rotée: row reste, col devient row inversé
    # Formule: excel_r = n_rows - r0 - h + 2, excel_c = c0 + 2
    # Taille rotée: hauteur_excel = h (inchangé), largeur_excel = w... non
    # CCW: (c, r) → (H-1-r, c), taille (w,h) → taille_excel (h, w)
    for (r0, c0, w, h, lbl, color) in placed:
        fill = PatternFill('solid', start_color=color)
        # Rotation CCW ¼ de tour depuis l'orientation de base (ligne 0 jeu en bas Excel).
        # Formule : new_r = n_cols - c0 - w + 1, new_c = n_rows - r0 - h + 2
        # Nouveau terrain : n_cols lignes × n_rows colonnes
        excel_r1 = n_cols - c0 - w + 1
        excel_r2 = n_cols - c0
        excel_c1 = n_rows - r0 - h + 2
        excel_c2 = n_rows - r0 + 1
        excel_r1 = max(2, excel_r1); excel_r2 = max(2, excel_r2)
        excel_c1 = max(2, excel_c1); excel_c2 = max(2, excel_c2)
        # Après rotation CCW: dimensions Excel = h_excel × w_excel
        # où h_excel = h_jeu (lignes), w_excel = h_jeu... 
        # Correction: excel_r = n_rows-r0-h+2..n_rows-r0+1, excel_c = c0+2..c0+h+1
        # Vérifie les bornes
        if excel_r1 < 2 or excel_c1 < 2 or excel_r2 > n_rows+1 or excel_c2 > n_cols+10:
            pass  # hors limites, skip silencieux
        elif excel_r1 != excel_r2 or excel_c1 != excel_c2:
            try:
                ws_t.merge_cells(
                    start_row=excel_r1, start_column=excel_c1,
                    end_row=excel_r2,   end_column=excel_c2
                )
            except Exception:
                pass  # chevauchement ignoré, on écrit quand même la cellule
        cell = ws_t.cell(excel_r1, excel_c1)
        from openpyxl.cell.cell import MergedCell as _MC
        if not isinstance(cell, _MC):
            cell.value     = lbl
            cell.font      = Font(name='Arial', size=7, bold=True)
            cell.alignment = ctr
            cell.fill      = fill
            cell.border    = Border(left=thick, right=thick, top=thick, bottom=thick)

    # Dimensions des colonnes et lignes (après CCW: n_rows colonnes, n_cols lignes)
    ws_t.column_dimensions['A'].width = 4
    for c in range(n_rows):
        ws_t.column_dimensions[get_column_letter(c + 2)].width = 4
    for r in range(n_cols + 1):
        ws_t.row_dimensions[r + 1].height = 20

    # Fige les volets (ligne 1 + colonne A)
    ws_t.freeze_panes = 'B2'

    # ── Onglet Batiments ──
    ws_b = wb.create_sheet('Batiments')
    headers = ['Nom', 'Ligne', 'Colonne', 'Hauteur', 'Largeur', 'Niveau',
               'Type', 'Culture', 'Rayonnement', 'Seuil_25%', 'Seuil_50%', 'Seuil_100%',
               'Nom_complet']

    for i, h_txt in enumerate(headers, 1):
        cell = ws_b.cell(1, i)
        cell.value = h_txt
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = ctr
        cell.border = brd_cell

    for row_idx, (_, b) in enumerate(rows_df.iterrows(), 2):
        if 'Largeur' in b and 'Hauteur' in b and pd.notna(b['Largeur']) and pd.notna(b['Hauteur']) and int(b['Largeur']) > 0:
            w, h = int(b['Largeur']), int(b['Hauteur'])
        else:
            bkey = get_key(b['Nom_complet'], DIMS)
            w, h = DIMS[bkey]
        cat  = get_cat(b['Nom_complet'])
        cult_raw = int(b['Culture'])     if 'Culture'     in b.index and pd.notna(b.get('Culture'))     else 0
        ray_raw  = int(b['Rayonnement']) if 'Rayonnement' in b.index and pd.notna(b.get('Rayonnement')) else 0
        # Culture/rayonnement : vrai seulement pour les bâtiments culturels
        # Les fermes/maisons/workshops/casernes ont cult=0, ray=0
        nom = b['Nom_complet']
        niveau_b = int(b['Niveau']) if 'Niveau' in b.index and pd.notna(b.get('Niveau')) else 1
        ere_idx = niveau_b  # f18 = index de l'ère dans le CityDTO (approximatif)
        if is_real_culture(nom, ray_raw):
            # Si culture=0 dans le CSV mais c'est un Evolving/DynamicAge → calcule depuis formule
            if cult_raw == 0 and any(k in nom for k in ['Evolving', 'DynamicAge']):
                cult = compute_evolving_culture(nom, niveau_b, ere_idx)
                ray  = compute_evolving_range(nom, niveau_b) if cult > 0 else ray_raw
            else:
                cult = cult_raw
                ray  = ray_raw
        else:
            cult, ray = 0, 0
        # Seuils de culture : depuis le CSV (colonnes Seuil25/50/100) ou DIMS_THRESHOLDS
        def _s(col_name, pct):
            v = b.get(col_name, '') if col_name in b.index else ''
            val = int(v) if pd.notna(v) and str(v) not in ('', 'None', '0') and int(float(str(v))) > 0 else None
            if val: return val
            return _get_threshold(b['Nom_complet'], pct)
        s25  = _s('Seuil25',  25)
        s50  = _s('Seuil50',  50)
        s100 = _s('Seuil100', 100)
        # Niveau : depuis le CSV (colonne Niveau = f18 du CityDTO) ou depuis le nom
        level = int(b['Niveau']) if 'Niveau' in b.index and pd.notna(b.get('Niveau')) and int(b.get('Niveau', 0)) > 0 else get_level(b['Nom_complet'])
        color = get_color(b['Nom_complet'], ray_raw)
        fill  = PatternFill('solid', start_color=color)

        row_data = [
            clean_name(b['Nom_complet']),
            int(b['Ligne']),
            int(b['Colonne']),
            h, w, level,
            cat, cult, ray,
            s25, s50, s100,
            b['Nom_complet'],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_b.cell(row_idx, col_idx)
            cell.value = val
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.alignment = lft if col_idx in (1, 6, 9, 10) else ctr
            cell.border = brd_cell

    col_widths = [35, 8, 10, 8, 8, 12, 10, 12, 18, 55]
    for i, cw in enumerate(col_widths, 1):
        ws_b.column_dimensions[get_column_letter(i)].width = cw

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Interface Streamlit ──
uploaded = st.file_uploader(
    "Uploadez le fichier CSV exporté depuis le jeu",
    type=["csv"],
    help="Fichier généré par le userscript RoC dans Safari sur iPad"
)

if uploaded:
    try:
        df = pd.read_csv(uploaded)
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].str.strip('"')

        # Filtre les valeurs aberrantes (uint32 overflow de coordonnées négatives)
        df['Ligne']   = pd.to_numeric(df['Ligne'],   errors='coerce')
        df['Colonne'] = pd.to_numeric(df['Colonne'], errors='coerce')
        df = df[
            (df['Ligne']   >= 0) & (df['Ligne']   < 10000) &
            (df['Colonne'] >= 0) & (df['Colonne'] < 10000)
        ].copy()

        st.success(f"✓ {len(df)} bâtiments chargés")

        villes = sorted(df['Ville'].unique().tolist())
        default_idx = villes.index('City_Capital') if 'City_Capital' in villes else 0
        ville_sel = st.selectbox("Choisissez la ville à exporter", villes, index=default_idx)

        df_ville = df[df['Ville'] == ville_sel]
        col1, col2, col3 = st.columns(3)
        col1.metric("Bâtiments", len(df_ville))
        col2.metric("Colonnes max", int(df_ville['Colonne'].max()))
        col3.metric("Lignes max",   int(df_ville['Ligne'].max()))

        with st.expander("Aperçu des bâtiments"):
            st.dataframe(
                df_ville[['Nom_complet', 'Ligne', 'Colonne']].reset_index(drop=True),
                height=300
            )

        buf = build_excel(df, ville_sel)
        if buf:
            st.download_button(
                label="⬇️ Télécharger le fichier Excel",
                data=buf,
                file_name=f"ville_{ville_sel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            st.info("Ce fichier Excel est prêt à être uploadé dans l'optimiseur de ville.")
        else:
            st.error("Aucun bâtiment trouvé pour cette ville.")

    except Exception as e:
        st.error(f"Erreur : {e}")
        st.exception(e)

else:
    st.info("👆 Uploadez le CSV pour commencer.")
    with st.expander("Comment obtenir le CSV ?"):
        st.markdown("""
1. Installez **RoC_export_ville.user.js** dans Userscripts sur iPad
2. Ouvrez Safari → `https://u0.riseofcultures.com/`
3. Connectez-vous et attendez que votre ville soit chargée
4. Un panneau vert apparaît en bas à droite
5. Tapez **⬇ Télécharger CSV**
6. Uploadez ce fichier ici
        """)
