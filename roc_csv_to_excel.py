"""
Convertit le CSV exporté par le userscript RoC en fichier Excel
compatible avec l'optimiseur de ville (onglets Terrain + Batiments).

Usage:
    python roc_csv_to_excel.py roc_ville_20260416_1800.csv
    python roc_csv_to_excel.py roc_ville_20260416_1800.csv --ville "City_Capital"
"""

import sys
import os
import csv
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Catalogue des dimensions de bâtiments RoC ──
# Format: nom_complet_ou_pattern → (largeur, hauteur)
# Les dimensions sont en cases de la grille.
BUILDING_DIMENSIONS = {
    # Culture Sites
    'CultureSite_Large':    (3, 3),
    'CultureSite_Moderate': (2, 2),
    'CultureSite_Compact':  (2, 1),
    'CultureSite_Small':    (1, 2),
    'CultureSite_Little':   (1, 1),
    'CultureSite_Luxurious':(2, 2),
    # Maisons
    'Home_Small':           (2, 2),
    'Home_Average':         (2, 2),
    'Home_Large':           (3, 2),
    'Home_Medium':          (2, 2),
    # Fermes
    'Farm_Rural':           (3, 2),
    'Farm_Domestic':        (2, 2),
    'Farm_Pastoral':        (3, 3),
    # Casernes
    'Barracks_Infantry':    (3, 3),
    'Barracks_Ranged':      (3, 3),
    'Barracks_Cavalry':     (3, 3),
    'Barracks_Siege':       (5, 6),
    # CityHall
    'CityHall':             (4, 4),
    # Évolutifs (taille typique)
    'Evolving':             (3, 3),
    # Collectables
    'Collectable':          (2, 2),
    # Wonders
    'Wonder':               (4, 4),
    # Irrigation (Arabia)
    'Irrigation_Noria':     (2, 2),
    'Irrigation_Large':     (3, 2),
    'Irrigation_Medium':    (2, 2),
    'Irrigation_Small':     (1, 2),
    'Irrigation_Oasis':     (3, 3),
    # Marchands
    'Merchant_Average':     (2, 2),
    # Ferme chameau
    'CamelFarm_Average':    (3, 3),
    # Workshop Arabia
    'Workshop_Coffee':      (2, 2),
    'Workshop_Incense':     (2, 2),
    'Workshop_OilLamp':     (2, 2),
    'Workshop_Carpet':      (2, 2),
    # Generic fallback
    'DEFAULT':              (2, 2),
}

# Types de bâtiments et leur catégorie pour l'optimiseur
BUILDING_CATEGORIES = {
    'CultureSite': 'Culturel',
    'Barracks':    'Caserne',
    'Farm':        'Production',
    'Home':        'Habitation',
    'CityHall':    'Neutre',
    'Evolving':    'Producteur',
    'Collectable': 'Neutre',
    'Wonder':      'Neutre',
    'Irrigation':  'Neutre',
    'Merchant':    'Production',
    'CamelFarm':   'Production',
    'Workshop':    'Production',
    'DEFAULT':     'Neutre',
}

# Valeur culture des sites culturels
CULTURE_VALUES = {
    'CultureSite_Large':     600,
    'CultureSite_Moderate':  350,
    'CultureSite_Compact':   200,
    'CultureSite_Small':     150,
    'CultureSite_Little':    100,
    'CultureSite_Luxurious': 700,
}

# Rayonnement des sites culturels
CULTURE_RANGE = {
    'CultureSite_Large':     3,
    'CultureSite_Moderate':  2,
    'CultureSite_Compact':   1,
    'CultureSite_Small':     2,
    'CultureSite_Little':    1,
    'CultureSite_Luxurious': 2,
}


def get_building_key(full_name):
    """Extrait la clé de dimensions depuis le nom complet."""
    # Cherche le pattern de type+variante dans le nom
    for key in BUILDING_DIMENSIONS:
        if key == 'DEFAULT':
            continue
        if key in full_name:
            return key
    return 'DEFAULT'


def get_category_key(full_name):
    """Extrait la catégorie depuis le nom complet."""
    for key in BUILDING_CATEGORIES:
        if key == 'DEFAULT':
            continue
        if key in full_name:
            return key
    return 'DEFAULT'


def get_display_name(full_name):
    """Génère un nom d'affichage court."""
    # Enlève "Building_" du début
    name = full_name.replace('Building_', '')
    # Enlève le niveau final (ex: "_1", "_2")
    parts = name.split('_')
    if parts and parts[-1].isdigit():
        parts = parts[:-1]
    # Garde les 3-4 dernières parties les plus descriptives
    if len(parts) > 4:
        parts = parts[-3:]
    return '_'.join(parts)


def parse_csv(csv_path):
    """Lit le CSV exporté par le userscript."""
    buildings = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            buildings.append({
                'city':     row.get('Ville', '').strip('"'),
                'name':     row.get('Nom_complet', '').strip('"'),
                'type':     row.get('Type', '').strip('"'),
                'era':      row.get('Ere', '').strip('"'),
                'variant':  row.get('Variante', '').strip('"'),
                'level':    int(row.get('Niveau', 1) or 1),
                'col':      int(row.get('Colonne', 0) or 0),
                'row':      int(row.get('Ligne', 0) or 0),
            })
    return buildings


def compute_terrain_size(buildings):
    """Calcule les dimensions minimales du terrain."""
    if not buildings:
        return 60, 60
    max_col = max(b['col'] for b in buildings)
    max_row = max(b['row'] for b in buildings)
    # Ajoute une marge de 5 cases
    return max(max_row + 5, 20), max(max_col + 5, 20)


def build_excel(buildings, output_path, ville_filter=None):
    """Génère le fichier Excel pour l'optimiseur."""

    # Filtre par ville si demandé
    if ville_filter:
        buildings = [b for b in buildings if b['city'] == ville_filter]
    
    if not buildings:
        print(f"Aucun bâtiment trouvé pour la ville '{ville_filter}'")
        return False

    n_rows, n_cols = compute_terrain_size(buildings)
    print(f"Ville: {buildings[0]['city']} - {len(buildings)} bâtiments")
    print(f"Terrain estimé: {n_rows} lignes × {n_cols} colonnes")

    wb = Workbook()

    # ── Styles ──
    hdr_fill  = PatternFill('solid', start_color='366092')
    hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell_font = Font(name='Arial', size=10)
    center    = Alignment(horizontal='center', vertical='center')
    left      = Alignment(horizontal='left', vertical='center')

    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Couleurs terrain
    FREE_COLOR  = 'FFFFFF'
    BORDER_COLOR = 'C0C0C0'
    
    # Couleurs bâtiments par catégorie
    CAT_COLORS = {
        'Culturel':    'FFE699',
        'Caserne':     'FF0000',
        'Production':  '92D050',
        'Habitation':  '00B0F0',
        'Producteur':  'FF66CC',
        'Neutre':      'D9D9D9',
    }

    # ── Onglet 1 : Terrain ──
    ws_terrain = wb.active
    ws_terrain.title = 'Terrain'

    # En-tête : numéros de colonnes (ligne 1)
    ws_terrain.cell(1, 1).value = ''
    for c in range(n_cols):
        cell = ws_terrain.cell(1, c + 2)
        cell.value = c
        cell.font = Font(bold=True, name='Arial', size=9)
        cell.alignment = center

    # Numéros de lignes (colonne A) + cases du terrain
    # Crée une grille avec les bâtiments placés
    grid = {}
    for b in buildings:
        bkey = get_building_key(b['name'])
        dims = BUILDING_DIMENSIONS.get(bkey, (2, 2))
        w, h = dims
        cat = BUILDING_CATEGORIES.get(get_category_key(b['name']), 'Neutre')
        short = get_display_name(b['name'])
        for dr in range(h):
            for dc in range(w):
                grid[(b['row'] + dr, b['col'] + dc)] = {
                    'cat': cat,
                    'label': short if (dr == 0 and dc == 0) else '',
                    'name': b['name'],
                }

    for r in range(n_rows):
        # Numéro de ligne
        cell_num = ws_terrain.cell(r + 2, 1)
        cell_num.value = r
        cell_num.font = Font(bold=True, name='Arial', size=9)
        cell_num.alignment = center

        for c in range(n_cols):
            cell = ws_terrain.cell(r + 2, c + 2)
            if (r, c) in grid:
                info = grid[(r, c)]
                cat_color = CAT_COLORS.get(info['cat'], 'D9D9D9')
                cell.fill = PatternFill('solid', start_color=cat_color)
                if info['label']:
                    cell.value = info['label']
                    cell.font = Font(name='Arial', size=7, bold=True)
                cell.alignment = center
                cell.border = border
            else:
                # Case libre
                cell.fill = PatternFill('solid', start_color=FREE_COLOR)
                cell.border = border

    # Largeur colonnes terrain
    ws_terrain.column_dimensions['A'].width = 5
    for c in range(n_cols):
        col_letter = get_column_letter(c + 2)
        ws_terrain.column_dimensions[col_letter].width = 4
    # Hauteur lignes
    ws_terrain.row_dimensions[1].height = 15
    for r in range(n_rows):
        ws_terrain.row_dimensions[r + 2].height = 15

    # ── Onglet 2 : Batiments ──
    ws_bld = wb.create_sheet('Batiments')
    
    headers = ['Nom', 'Ligne', 'Colonne', 'Hauteur', 'Largeur',
               'Type', 'Culture', 'Rayonnement', 'Ere', 'Nom_complet']
    
    for i, h in enumerate(headers, 1):
        cell = ws_bld.cell(1, i)
        cell.value = h
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    for row_idx, b in enumerate(buildings, 2):
        bkey  = get_building_key(b['name'])
        dims  = BUILDING_DIMENSIONS.get(bkey, (2, 2))
        w, h  = dims
        catk  = get_category_key(b['name'])
        cat   = BUILDING_CATEGORIES.get(catk, 'Neutre')
        cult  = CULTURE_VALUES.get(bkey, 0)
        ray   = CULTURE_RANGE.get(bkey, 0)
        short = get_display_name(b['name'])

        row_data = [
            short,          # Nom
            b['row'],       # Ligne
            b['col'],       # Colonne
            h,              # Hauteur
            w,              # Largeur
            cat,            # Type
            cult,           # Culture
            ray,            # Rayonnement
            b['era'],       # Ere
            b['name'],      # Nom_complet
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_bld.cell(row_idx, col_idx)
            cell.value = val
            cell.font = cell_font
            cell.alignment = left if col_idx in (1, 6, 9, 10) else center
            cell.border = border

        # Couleur par catégorie
        cat_color = CAT_COLORS.get(cat, 'D9D9D9')
        for col_idx in range(1, len(headers) + 1):
            ws_bld.cell(row_idx, col_idx).fill = PatternFill('solid', start_color=cat_color)

    # Largeur colonnes bâtiments
    col_widths = [35, 8, 10, 8, 8, 12, 10, 12, 18, 55]
    for i, w in enumerate(col_widths, 1):
        ws_bld.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"✓ Fichier Excel créé : {output_path}")
    print(f"  - Onglet Terrain  : {n_rows} × {n_cols} cases")
    print(f"  - Onglet Batiments: {len(buildings)} bâtiments")
    return True


def list_cities(csv_path):
    """Liste les villes disponibles dans le CSV."""
    buildings = parse_csv(csv_path)
    cities = {}
    for b in buildings:
        cities[b['city']] = cities.get(b['city'], 0) + 1
    print("Villes disponibles :")
    for city, count in sorted(cities.items()):
        print(f"  {city}: {count} bâtiments")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convertit le CSV RoC en Excel pour l\'optimiseur')
    parser.add_argument('csv_file', help='Fichier CSV exporté par le userscript')
    parser.add_argument('--ville', default=None, help='Nom de la ville à exporter (ex: City_Capital)')
    parser.add_argument('--list', action='store_true', help='Liste les villes disponibles')
    parser.add_argument('--output', default=None, help='Nom du fichier de sortie')
    args = parser.parse_args()

    if not os.path.exists(args.csv_file):
        print(f"Erreur: fichier '{args.csv_file}' introuvable")
        sys.exit(1)

    if args.list:
        list_cities(args.csv_file)
        sys.exit(0)

    buildings = parse_csv(args.csv_file)

    # Si pas de filtre de ville, prend City_Capital par défaut
    if not args.ville:
        cities_found = list(set(b['city'] for b in buildings))
        if 'City_Capital' in cities_found:
            args.ville = 'City_Capital'
            print(f"Ville par défaut: City_Capital")
        elif cities_found:
            args.ville = cities_found[0]
            print(f"Ville sélectionnée: {args.ville}")

    # Nom de sortie
    if not args.output:
        base = os.path.splitext(args.csv_file)[0]
        ville_slug = (args.ville or 'ville').replace(' ', '_').replace('/', '_')
        args.output = f"{base}_{ville_slug}.xlsx"

    build_excel(buildings, args.output, args.ville)
